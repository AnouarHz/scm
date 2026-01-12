from __future__ import annotations

from pathlib import Path
import os, time, json, base64, hmac, hashlib, re
from io import BytesIO
from dataclasses import dataclass
from typing import Dict, Any, List, Tuple, Optional

import openpyxl
from fastapi import FastAPI, Request, Form, HTTPException, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, Response

import smtplib
from email.message import EmailMessage

from starlette.templating import Jinja2Templates

# ==========================================================
# PATHS + CONFIG
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent
ROOT = BASE_DIR / "ROOT"
ROOT.mkdir(exist_ok=True)

SCALE_XLSX = ROOT / "3. SCM-Anwendungen(MA)_Gesamtbewertung.xlsx"
SUBMISSIONS_DIR = ROOT / "Form_Submissions"
SUBMISSIONS_DIR.mkdir(exist_ok=True)

APP_SECRET = os.getenv("SCM_FORM_SECRET", "CHANGE_ME_SECRET")
BASE_URL = os.getenv("SCM_FORM_BASE_URL", "http://localhost:8000")
TOKEN_TTL = int(os.getenv("SCM_TOKEN_TTL_SECONDS", str(7 * 24 * 3600)))
SCALE_SHEET = os.getenv("SCM_SCALE_SHEET", "Skala")

SOURCE_COL_NAME = "Datenherkunft / Bewertung"
SOURCE_MANUAL_VALUE = "Manuelle Bewertung"

CACHE_TTL = 30

# ==========================================================
# SMTP / MAIL CONFIG (Aufgabe 3)
# ==========================================================
MAIL_ENABLED = os.getenv("SCM_MAIL_ENABLED", "1").strip() not in ("0", "false", "False", "no", "NO")
MAIL_TO = os.getenv("SCM_MAIL_TO", "anouar.hizaoui001@stud.fh-dortmund.de").strip()

SMTP_HOST = os.getenv("SCM_SMTP_HOST", "mail.gmx.net").strip()
SMTP_PORT = int(os.getenv("SCM_SMTP_PORT", "587"))
SMTP_USER = os.getenv("SCM_SMTP_USER", "scmuser@gmx.de").strip()
SMTP_PASS = os.getenv("SCM_SMTP_PASS", "scmuser1!").strip()
SMTP_FROM = os.getenv("SCM_SMTP_FROM", SMTP_USER).strip()  # fallback

SMTP_USE_STARTTLS = os.getenv("SCM_SMTP_STARTTLS", "1").strip() not in ("0", "false", "False", "no", "NO")

app = FastAPI(title="SCM Form Backend")

templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


# ==========================================================
# SMALL HELPERS
# ==========================================================
def norm(s: Any) -> str:
    s = "" if s is None else str(s)
    s = s.replace("\u00A0", " ").replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def clean(s: Any) -> str:
    return "" if s is None else str(s).strip()

def sanitize_filename_part(s: Any) -> str:
    return re.sub(r"[^\w\-]", "_", clean(s))

def submission_path(round_id: str, supplier_id: str) -> Path:
    return SUBMISSIONS_DIR / f"submission_{sanitize_filename_part(supplier_id)}_R{sanitize_filename_part(round_id)}.json"


# ==========================================================
# TOKEN (HMAC)
# ==========================================================
def _b64e(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).decode("utf-8").rstrip("=")

def _b64d(s: str) -> bytes:
    return base64.urlsafe_b64decode(s + "==")

def make_token(payload: Dict[str, Any]) -> str:
    raw = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    sig = hmac.new(APP_SECRET.encode("utf-8"), raw, hashlib.sha256).digest()
    return f"{_b64e(raw)}.{_b64e(sig)}"

def read_token(token: str) -> Dict[str, Any]:
    try:
        part_raw, part_sig = token.split(".", 1)
        raw = _b64d(part_raw)
        sig = _b64d(part_sig)

        exp_sig = hmac.new(APP_SECRET.encode("utf-8"), raw, hashlib.sha256).digest()
        if not hmac.compare_digest(sig, exp_sig):
            raise ValueError("bad signature")

        payload = json.loads(raw.decode("utf-8"))
        exp = payload.get("exp")
        if exp is not None and time.time() > float(exp):
            raise ValueError("expired")

        if not payload.get("supplier_id") or not payload.get("round_id"):
            raise ValueError("missing supplier_id/round_id")

        return payload
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"invalid token: {e}")


# ==========================================================
# MAIL SENDER (Aufgabe 3)
# ==========================================================
def send_submission_mail(round_id: str, supplier_id: str, answers: Dict[str, int]) -> None:
    """
    Sends an email with the XLSX submission attached.
    Runs as BackgroundTask to keep the submit endpoint fast.
    """
    if not MAIL_ENABLED:
        return

    # Basic config validation (silent fail -> only server log)
    if not (SMTP_HOST and SMTP_PORT and SMTP_USER and SMTP_PASS and SMTP_FROM and MAIL_TO):
        print("[MAIL] SMTP config missing (SCM_SMTP_HOST/PORT/USER/PASS/FROM or SCM_MAIL_TO). Mail not sent.")
        return

    xlsx_bytes = build_reply_xlsx(answers)
    filename = f"Antwort_{supplier_id}_R{round_id}.xlsx"

    msg = EmailMessage()
    msg["Subject"] = f"SCM Bewertung eingegangen | Runde {round_id} | {supplier_id}"
    msg["From"] = SMTP_FROM
    msg["To"] = MAIL_TO
    msg.set_content(
        "Hallo,\n\n"
        f"es ist eine neue Bewertung eingegangen.\n\n"
        f"Runde: {round_id}\n"
        f"Lieferant: {supplier_id}\n\n"
        "Im Anhang finden Sie die exportierte Antwortdatei (XLSX).\n\n"
        "Viele Grüße\n"
        "SCM Formular-Service\n"
    )

    msg.add_attachment(
        xlsx_bytes,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as smtp:
            smtp.ehlo()
            if SMTP_USE_STARTTLS:
                smtp.starttls()
                smtp.ehlo()
            smtp.login(SMTP_USER, SMTP_PASS)
            smtp.send_message(msg)
        print(f"[MAIL] Sent submission mail to {MAIL_TO} ({filename}).")
    except Exception as e:
        print(f"[MAIL] Failed to send mail: {e}")


# ==========================================================
# SCALE + MANUAL CRITERIA (single pass openpyxl)
# ==========================================================
def _find_header(ws, header_contains: str) -> Tuple[int, int]:
    tgt = norm(header_contains)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            if tgt in norm(cell.value):
                return cell.row, cell.column
    raise HTTPException(500, f"Header '{header_contains}' nicht gefunden (Sheet '{ws.title}').")

def _find_col_in_row(ws, row: int, needle_contains: str) -> Optional[int]:
    tgt = norm(needle_contains)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row, c).value
        if v is not None and tgt in norm(v):
            return c
    return None

def load_scale_and_manual() -> Tuple[Dict[str, Dict[int, str]], List[str]]:
    if not SCALE_XLSX.exists():
        raise HTTPException(500, f"Excel nicht gefunden: {SCALE_XLSX}")

    wb = openpyxl.load_workbook(SCALE_XLSX, data_only=True)
    if SCALE_SHEET not in wb.sheetnames:
        raise HTTPException(500, f"Sheet '{SCALE_SHEET}' nicht gefunden in Excel.")

    ws = wb[SCALE_SHEET]

    header_row, src_col = _find_header(ws, SOURCE_COL_NAME)
    crit_col = _find_col_in_row(ws, header_row, "kriter") or 1

    scale_cols = {0: 5, 20: 6, 40: 7, 60: 8, 80: 9, 100: 10}

    scale: Dict[str, Dict[int, str]] = {}
    manual: List[str] = []

    manual_marker = norm(SOURCE_MANUAL_VALUE)

    for r in range(header_row + 1, ws.max_row + 1):
        crit = clean(ws.cell(r, crit_col).value)
        if not crit or norm(crit) in ("nan", "none"):
            continue

        scale[crit] = {pts: clean(ws.cell(r, col).value) for pts, col in scale_cols.items()}

        src_val = ws.cell(r, src_col).value
        if norm(src_val) == manual_marker:
            manual.append(crit)

    if not scale:
        raise HTTPException(500, "Skala konnte nicht geladen werden (keine Kriterien gefunden).")

    if not manual:
        raise HTTPException(
            500,
            f"Header gefunden (Row {header_row}, Col {src_col}), aber keine Zeilen mit '{SOURCE_MANUAL_VALUE}'."
        )

    manual = list(dict.fromkeys(manual))
    return scale, manual


@dataclass
class _Cache:
    ts: float = 0.0
    scale: Optional[Dict[str, Dict[int, str]]] = None
    manual: Optional[List[str]] = None

_CACHE = _Cache()

def get_scale_and_manual() -> Tuple[Dict[str, Dict[int, str]], List[str]]:
    now = time.time()
    if _CACHE.scale is not None and (now - _CACHE.ts) < CACHE_TTL:
        return _CACHE.scale, _CACHE.manual or []
    scale, manual = load_scale_and_manual()
    _CACHE.ts, _CACHE.scale, _CACHE.manual = now, scale, manual
    return scale, manual


# ==========================================================
# XLSX Builder (für Bot + Mail)
# ==========================================================
def build_reply_xlsx(answers: Dict[str, int]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Antwort"
    ws.append(["Kriterium", "Bewertung"])
    for k, v in answers.items():
        ws.append([k, int(v)])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ==========================================================
# Storage
# ==========================================================
def save_submission(round_id: str, supplier_id: str, answers: Dict[str, int]) -> Dict[str, Any]:
    data = {"round_id": round_id, "supplier_id": supplier_id, "submitted_at": time.time(), "answers": answers}
    submission_path(round_id, supplier_id).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data

def load_submission(round_id: str, supplier_id: str) -> Dict[str, Any]:
    p = submission_path(round_id, supplier_id)
    if not p.exists():
        raise HTTPException(404, "submission not found")
    return json.loads(p.read_text(encoding="utf-8"))


# ==========================================================
# Routes
# ==========================================================
@app.get("/", response_class=HTMLResponse)
async def home():
    scale, manual = get_scale_and_manual()
    return HTMLResponse(
        f"""
        <h3>SCM Formular-Service ✅</h3>
        <p>Excel: <code>{SCALE_XLSX}</code></p>
        <p>Manuelle Kriterien (nach '{SOURCE_COL_NAME}' = '{SOURCE_MANUAL_VALUE}'): <b>{len(manual)}</b></p>
        <p>Test: <code>{BASE_URL}/issue-link?supplier_id=K_1&round_id=12345678</code></p>
        """
    )

@app.get("/issue-link")
async def issue_link(supplier_id: str, round_id: str):
    supplier_id, round_id = clean(supplier_id), clean(round_id)
    if not supplier_id or not round_id:
        raise HTTPException(400, "missing supplier_id or round_id")

    token = make_token({"supplier_id": supplier_id, "round_id": round_id, "exp": time.time() + TOKEN_TTL})
    url = f"{BASE_URL.rstrip('/')}/evaluate?token={token}"
    return JSONResponse({"url": url, "token": token})

@app.get("/evaluate", response_class=HTMLResponse)
async def evaluate(request: Request, token: str):
    payload = read_token(token)
    scale, manual = get_scale_and_manual()

    items = [{"crit": c, "scale": scale[c]} for c in manual]
    return templates.TemplateResponse(
        "form.html",
        {
            "request": request,
            "token": token,
            "round_id": payload["round_id"],
            "supplier_id": payload["supplier_id"],
            "items": list(enumerate(items)),
        },
    )

@app.post("/submit", response_class=HTMLResponse)
async def submit(
    request: Request,
    background_tasks: BackgroundTasks,
    token: str = Form(...),
):
    payload = read_token(token)
    scale, manual = get_scale_and_manual()

    form = await request.form()
    answers: Dict[str, int] = {}

    for i, crit in enumerate(manual):
        val = form.get(f"c_{i}")
        if val is None:
            raise HTTPException(400, f"missing answer for {crit}")
        pts = int(val)
        if pts not in (0, 20, 40, 60, 80, 100):
            raise HTTPException(400, f"invalid scale value: {pts}")
        answers[crit] = pts

    save_submission(payload["round_id"], payload["supplier_id"], answers)

    # ✅ Aufgabe 3: Mail automatisch senden (asynchron im Hintergrund)
    background_tasks.add_task(send_submission_mail, payload["round_id"], payload["supplier_id"], answers)

    return HTMLResponse(
        f"""
        <h3>Danke! ✅</h3>
        <p>Ihre Bewertung für <b>{payload["supplier_id"]}</b> (Runde <b>{payload["round_id"]}</b>) wurde gespeichert.</p>
        <p>Sie können dieses Fenster jetzt schließen.</p>
        """
    )

@app.get("/api/submissions")
async def api_submissions(round_id: str):
    rid = clean(round_id)
    out = []
    for f in SUBMISSIONS_DIR.glob("submission_*_R*.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            if str(data.get("round_id")) == rid:
                out.append({"supplier_id": data.get("supplier_id"), "submitted_at": float(data.get("submitted_at", 0))})
        except Exception:
            continue
    out.sort(key=lambda x: x["submitted_at"])
    return JSONResponse(out)

@app.get("/api/xlsx")
async def api_xlsx(round_id: str, supplier_id: str):
    data = load_submission(round_id, supplier_id)
    answers = {k: int(v) for k, v in (data.get("answers") or {}).items()}
    xlsx_bytes = build_reply_xlsx(answers)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Antwort_{supplier_id}_R{round_id}.xlsx"'},
    )
